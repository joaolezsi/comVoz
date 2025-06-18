from flask import Blueprint, request, jsonify
from werkzeug.utils import secure_filename
import os
from openpyxl import load_workbook
import json

excel_bp = Blueprint('excel', __name__)

# Configuração para upload de arquivos
UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@excel_bp.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'error': 'Nenhum arquivo enviado'}), 400
    
    file = request.files['file']
    
    if file.filename == '':
        return jsonify({'error': 'Nenhum arquivo selecionado'}), 400
    
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        filepath = os.path.join(UPLOAD_FOLDER, filename)
        file.save(filepath)
        
        try:
            # Carrega o arquivo Excel
            wb = load_workbook(filepath)
            sheet = wb.active
            
            # Converte para JSON
            data = []
            headers = [cell.value for cell in sheet[1]]
            
            for row in sheet.iter_rows(min_row=2):
                row_data = {}
                for header, cell in zip(headers, row):
                    row_data[header] = cell.value
                data.append(row_data)
            
            # Remove o arquivo após processamento
            os.remove(filepath)
            
            return jsonify({
                'message': 'Arquivo processado com sucesso',
                'data': data
            }), 200
            
        except Exception as e:
            return jsonify({'error': f'Erro ao processar arquivo: {str(e)}'}), 500
    
    return jsonify({'error': 'Tipo de arquivo não permitido'}), 400 